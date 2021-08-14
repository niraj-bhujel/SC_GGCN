#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Sep 18 00:22:13 2020

@author: dl-asoro
"""
import sys
import os
import csv
import torch
import numpy as np
from openpyxl import load_workbook, Workbook
from collections import OrderedDict
import matplotlib.pyplot as plt
import cv2
import glob

def get_gpu_memory(acceptable_memory=64):
    import subprocess as sp
    
    free_memory = sp.check_output(['nvidia-smi', '--query-gpu=memory.free',
                                   '--format=csv,nounits,noheader'])
    total_memory = sp.check_output(['nvidia-smi', '--query-gpu=memory.total',
                                    '--format=csv,nounits,noheader'])
    
    free_memory = [int(x) for x in free_memory.decode('ascii').strip().split('\n')]
    total_memory = [int(x) for x in total_memory.decode('ascii').strip().split('\n')]
    
    for i in range(len(free_memory)):
        print('GPU {} with Total:{} MiB, Free:{} MiB'.format(i, total_memory[i], free_memory[i]))
        
    # memory_map = dict(zip(range(len(free_memory)), free_memory))
    return free_memory

def setup_gpu(gpu_id=None, memory=None):
    #setup gpu either by gpu_id or memory or both
    
    os.environ["CUDA_DEVICE_ORDER"] = "PCI_BUS_ID"
    
    gpu_memory = get_gpu_memory()
    ngpus = list(range(len(gpu_memory)))
    
    if len(ngpus)>0:
        #setup device by gpu_id and memory
        if gpu_id is not None and memory is not None:
            # available_gpus = [i for i, m in enumerate(gpu_memory) if m>memory]
            if gpu_memory[gpu_id]>memory:
                device = torch.device("cuda:{}".format(gpu_id))
                print('Using GPU {}'.format(gpu_id))
            else:
                raise Exception("GPU {} doesn't have required memory of {} MiB".format(gpu_id, memory))
        #setup device by gpu_id
        elif gpu_id is not None and memory is None:
            try:
                device = torch.device("cuda:{}".format(ngpus[gpu_id]))
                print('Using GPU {}'.format(ngpus[gpu_id]))
            except:
                raise Exception("GPU {} not available. Available idx are {}".format(gpu_id, ngpus))
        #setup device by memory only
        else:
            available_gpus = [i for i, m in enumerate(gpu_memory) if m>memory]
            try:
                device = torch.device("cuda:{}".format(available_gpus[0]))
                print('Using GPU {}'.format(available_gpus[0], len(ngpus)))
            except:
                raise Exception('GPUs not available with free memory {}'.format(memory))
    else:
        print('Using CPU..')
        device = torch.device("cpu")
        
    return device

def create_new_dir(new_dir):
    if not os.path.exists(new_dir):
        os.makedirs(new_dir)
    return new_dir

def standardize(data, mean=[-0.028, -0.015], sigma=[0.39, 0.34]):
    # data: node features (K, seq_len*2)
    _, T = data.shape
    T = int(T/2)
    data[:, :T] = (data[:, :T] - mean[0])/sigma[0]
    data[:, T:] = (data[:, T:] - mean[1])/sigma[1]
    return data

def save_ckpt(model, optimizer, scheduler, epoch, save_ckpt_dir):
    state = {
        'last_epoch': epoch,
        'state_dict': model.state_dict(),
        'optimizer': optimizer.state_dict(),
        'scheduler':scheduler.state_dict(),
    }
    torch.save(state, save_ckpt_dir+'model_epoch_{}.pth'.format(epoch))

def load_ckpt(model, optimizer, scheduler, epoch, load_ckpt_dir):
    checkpoint = torch.load(load_ckpt_dir + 'model_epoch_{}.pth'.format(epoch))
    model.load_state_dict(checkpoint['state_dict'])
    optimizer.load_state_dict(checkpoint['optimizer'])
    scheduler.load_state_dict(checkpoint['scheduler'])
    return model, optimizer, scheduler

def model_attributes(model, verbose=0):
    attributes = {k:v for k, v in model.__dict__.items() if not k.startswith('_')}
    
    if verbose>0:
        print(sorted(attributes.items()))
        
    return attributes

def model_parameters(model, verbose=0):
    if verbose>0:
        print('{:<30} {:<10} {:}'.format('Parame Name', 'Total Param', 'Param Shape'))
    total_params=0
    for name, param in model.named_parameters():
        if param.requires_grad:
            if verbose>0:
                print('{:<30} {:<10} {:}'.format(name, param.numel(), tuple(param.shape)))
            total_params+=param.numel()
    print('Total Trainable Parameters :{:<10}'.format(total_params))
    return total_params

def save_history(history, save_dir):
    # print('History saved to:', save_dir)
    row_head = list(history.keys())
    rows_val = np.around(np.array([val for key, val in history.items()]).T, 6)
    # row_num = rows_val[0].keys() #epoch number
    with open(save_dir+'/history.txt', 'a') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(row_head)
        writer.writerows(rows_val)
    csvfile.close()
    
def save_to_excel(file_path, row_data, header):
    # file_path = './out/GatedGCN/' + 'log_trial_results.xlsx'
    # Confirm file exists. 
    # If not, create it, add headers, then append new data
    try:
        wb = load_workbook(file_path)
        ws = wb.worksheets[0]  # select first worksheet
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(header) #header row
    ws.append(row_data)
    wb.save(file_path)
    wb.close()

def video_from_images(img_dir, dest_dir='./videos', fname ='video.mp4', frame_rate=2.5):
    if not os.path.exists(dest_dir):
        os.mkdir(dest_dir)
    img_list = []
    for filename in sorted(glob.glob(img_dir + '*.jpeg')):
        img = cv2.imread(filename)
        height, width, layers = img.shape
        size = (width, height)
        img_list.append(img)
    if len(img_list)>0:
        out = cv2.VideoWriter(fname, cv2.VideoWriter_fourcc(*'mp4v'), frame_rate, size)
        for img in img_list:
            out.write(img)
        out.release()