#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jul 14 16:18:54 2020

@author: niraj
"""
import math
import os
import numpy as np
from utils import read_file

data_set = 'eth'
data_dir = data_set = './datasets/' + data_set + '/' + 'val/'

obs_len=8
pred_len=12
seq_len = obs_len + pred_len
skip=1
max_peds_in_frame = 0
threshold = 0.002
min_ped = 1
feats_last=False

all_files = os.listdir(data_dir)
all_files = [os.path.join(data_dir, _path) for _path in all_files]
# print(all_files)

num_peds_in_seq = []
frames_list = []
id_list  = []
seq_list = []
seq_list_rel = []
loss_mask_list = []
non_linear_ped = []
all_data_by_id = {}

all_unique_peds = 0
for path in all_files:
    print(path)
    seq_name = os.path.basename(path)[:-4]
    data = read_file(path, delim='\t')
    
    # data_by_id = {}
    # for frame_id, person_id, x, y in data:
    #     if person_id not in data_by_id.keys():
    #         data_by_id[person_id] = []
    #     data_by_id[person_id].append([person_id, frame_id, x, y])
    # all_data_by_id[seq_name] = data_by_id
    
    #map the id to a new id
    unique_ids = np.unique(data[:, 1])
    unique_ids_new = all_unique_peds + np.arange(len(unique_ids))
    id_map = {pid:pid_new for pid, pid_new in zip(unique_ids, unique_ids_new)}
    all_unique_peds += len(unique_ids)
    #update id for each row
    data[:, 1] = np.array([id_map[pid] for pid in data[:, 1]])
    
    

    frames = np.unique(data[:, 0]).tolist()
    frame_data = []
    for frame in frames:
        frame_data.append(data[frame == data[:, 0], :])

    num_sequences = int(math.ceil((len(frames) - seq_len + 1) / skip))
    
    for idx in range(0, num_sequences * skip + 1, skip):
        
        curr_seq_data = np.concatenate(frame_data[idx:idx + seq_len], axis=0)

        peds_in_curr_seq = np.unique(curr_seq_data[:, 1])
        
        max_peds_in_frame = max(max_peds_in_frame, len(peds_in_curr_seq))
        
        # curr_seq_rel = np.zeros((len(peds_in_curr_seq), 2, seq_len)) #[num_peds, 2, seq_len]
        # curr_seq = np.zeros((len(peds_in_curr_seq), 2, seq_len))
        # curr_loss_mask = np.zeros((len(peds_in_curr_seq), seq_len))
        
        
        num_peds_considered = 0
        _non_linear_ped = []
        curr_frame_list = []
        curr_id_list = []
        curr_seq_list = []
        
        curr_seq_rel_list = []
        curr_loss_mask_list = []
        
        for _, ped_id in enumerate(peds_in_curr_seq):
            
            curr_ped_seq = curr_seq_data[curr_seq_data[:, 1] ==ped_id, :] #(20, 4)
            curr_ped_seq = np.around(curr_ped_seq, decimals=4)
            
            pad_front = frames.index(curr_ped_seq[0, 0]) - idx
            pad_end = frames.index(curr_ped_seq[-1, 0]) - idx + 1
            
            if pad_end - pad_front != seq_len:
                continue
            
            ped_frames = curr_ped_seq[:, 0]
            ped_seq = np.transpose(curr_ped_seq[:, 2:]) #(2, 20)
            
            # Make coordinates relative
            rel_ped_seq = np.zeros(ped_seq.shape)
            rel_ped_seq[:, 1:] = ped_seq[:, 1:] - ped_seq[:, :-1]
                         
            # _idx = num_peds_considered
            
            # curr_seq[_idx, :, pad_front:pad_end] = ped_seq
            # curr_seq_rel[_idx, :, pad_front:pad_end] = rel_ped_seq
            curr_frame_list.append(ped_frames)
            curr_id_list.append(ped_id)
            curr_seq_list.append(ped_seq)
            curr_seq_rel_list.append(rel_ped_seq)
            
            # Linear vs Non-Linear Trajectory
            # _non_linear_ped.append(poly_fit(ped_seq, pred_len, threshold))
            
            # curr_loss_mask[_idx, pad_front:pad_end] = 1
          
            # curr_loss_mask_list.append(np.ones(seq_len))
            
            num_peds_considered += 1
            
            # break
            
        if num_peds_considered > min_ped:
            
            # non_linear_ped += _non_linear_ped
            # loss_mask_list.append(curr_loss_mask_list)
            
            num_peds_in_seq.append(num_peds_considered)
            frames_list.append(curr_frame_list)
            id_list.append(curr_id_list)                        
            seq_list.append(curr_seq_list)
            
            seq_list_rel.append(curr_seq_rel_list)                
            
    # break


num_seq = len(seq_list)
frames_list_concat = np.concatenate(frames_list, axis=0) #[num_peds, 20]
id_list_concat = np.concatenate(id_list, axis=0)
seq_list_concat = np.concatenate(seq_list, axis=0)
seq_list_rel_concat = np.concatenate(seq_list_rel, axis=0)

# loss_mask_list_concat = np.concatenate(loss_mask_list, axis=0)
# non_linear_ped = np.asarray(non_linear_ped)


# Convert numpy -> Torch Tensor
obs_traj = seq_list_concat[:, :, :obs_len]
pred_traj = seq_list_concat[:, :, obs_len:]

obs_traj_rel = seq_list_rel_concat[:, :, :obs_len]
pred_traj_rel = seq_list_rel_concat[:, :, obs_len:]

obsv_frames = frames_list_concat[:, :obs_len]
target_frames = frames_list_concat[:, obs_len:]
# loss_mask = loss_mask_list_concat
# non_linear_ped = non_linear_ped

cum_start_idx = [0] + np.cumsum(num_peds_in_seq).tolist()
seq_start_end = [(start, end) for start, end in zip(cum_start_idx, cum_start_idx[1:])]

for start, end in seq_start_end:
    seq = obs_traj[start:end, :]
    seq_rel = obs_traj_rel[start:end, :]
    frames = obsv_frames[start:end, :]
    ped_ids = id_list_concat[start:end, np.newaxis]
    break

x_max, x_min = np.max(seq_list_concat[:, 0, :]), np.min(seq_list_concat[:, 0, :])
y_max, y_min = np.max(seq_list_concat[:, 1, :]), np.min(seq_list_concat[:, 1, :])
print('x_max:', x_max, 'x_min:', x_min, 'y_max:', y_max, 'y_min:', y_min)


#%%prepare dgl dataset
from scipy.spatial.distance import pdist, squareform
import dgl 

pbar = tqdm(total=len(seq_start_end))

num_neighbors = 1
for ss in range(len(seq_start_end)): #frame
    pbar.update(1)
    
    start, end = seq_start_end[ss] #number of peds per frame
    
    seq = obs_traj[start:end,:] #[num_peds, 2, 8]
    seq_rel = obs_traj_rel[start:end, :]
    
    seq_len = seq_.shape[2]
    
    num_nodes = seq_.shape[0]
    nodes_feats = np.reshape(seq_rel, (num_nodes, -1))
    
    nodes_dist = np.zeros((seq_len, num_nodes, num_nodes))
    for t in range(seq_len):
        nodes_coord = seq[:, :, t]
        # Compute distance matrix
        W_val = squareform(pdist(nodes_coord, metric='euclidean'))
        
        # Determine k-nearest neighbors for each node
        # knns = np.argpartition(W_val, kth=num_neighbors, axis=-1)[:, num_neighbors::-1]            
        nodes_dist[t, :, :] = W_val
    
    # Construct the DGL graph
    g = dgl.DGLGraph()
    g.add_nodes(num_nodes)
    g.ndata['feat'] = torch.Tensor(nodes_feats)
    
    
    edge_feats = []
    for i in range(num_nodes):
        for j in range(num_nodes):
            if i!=j:
                g.add_edge(i, j)
                edge_feats.append(nodes_dist[:, i, j])
    
    assert len(edge_feats) == g.number_of_edges()
    
    # Add edge features
    g.edata['feat'] = torch.Tensor(edge_feats).unsqueeze(-1)
    
    graph_lists.append(g)
    
    break

#%%
import math
import os
import numpy as np
from utils import read_file
import matplotlib.pyplot as plt
data_set = 'univ'
all_data = {}

for phase in ['train', 'val']:
    # phase = 'val'
    data_dir = './datasets/' + data_set + '/' + phase + '/'
    
    all_files = os.listdir(data_dir)
    all_files = [os.path.join(data_dir, _path) for _path in all_files]
    print(all_files)
    
    for file in all_files:
        seq_name = os.path.basename(file)[:-4]
        data = read_file(file, delim='\t')
        all_data[seq_name]=data
        
#%%

train_data_eth = all_data['biwi_eth_train']
train_peds_eth = np.unique(train_data_eth[:, 1])

val_data_eth = all_data['biwi_eth_val']
val_peds_eth = np.unique(val_data_eth[:, 1])

#some val_ped_ids appears in train_peds_id
common_peds = val_peds_eth[np.in1d(val_peds_eth, train_peds_eth)]
cmap = np.random.rand(len(common_peds), 3)
#verify trajectory in train and val belong to the same peds
for i in range(len(common_peds)):
    ped = common_peds[i]
    train_traj = train_data_eth[train_data_eth[:, 1]==ped, 2:4]
    val_traj = val_data_eth[val_data_eth[:, 1]==ped, 2:4]
    
    #plot train_traj and val_traj
    plt.plot(train_traj[:, 0], train_traj[:, 1], color=cmap[i], label='train')
    plt.plot(val_traj[:, 0], val_traj[:, 1], color=cmap[i], label='val')
    plt.legend()

#%%
#plot ground truth vs predicted path
import matplotlib.pyplot as plt

top_k = 3

traj_dir = os.path.join('./trajectories', os.path.basename(exp_path))
if not os.path.exists(traj_dir):
    os.makedirs(traj_dir)
cmap = np.random.random((num_of_objs, 3))
fig, ax = plt.subplots(figsize=(19, 10))
for ped in range(num_of_objs):
    y_trgt = np.array(raw_data_dict[step]['trgt'])[:, ped, :] #[12, 2]
    y_pred = np.array(raw_data_dict[step]['pred'])[:, :, ped, :] #[20, 12, 2]

    
    top_k_predIdx = np.argsort(fde_ls[ped])[:top_k]
    top_k_pred = y_pred[top_k_predIdx, :, :]
    
    ax.plot(y_trgt[:, 0], y_trgt[:, 1], linestyle='-', marker='*', label='gt', color=cmap[ped])
    for k in range(top_k):
        ax.plot(top_k_pred[k, :, 0,], top_k_pred[k, :, 1], linestyle='--', marker='+', label='pred_top_%d'%(k+1), color=cmap[ped])

plt.savefig(traj_dir + '/step_%d.jpg'%step)
# plt.legend()

#%%

min_x, max_x, min_y, max_y = (-0.3578, 15.4629, -0.1909, 13.2621)
def get_goal(position, min_x, max_x, min_y, max_y, n_cells_x, n_cells_y):
    """Divides the scene rectangle into a grid of cells and finds the cell idx where the current position falls"""
    x = position[0]
    y = position[1]

    x_steps = np.linspace(min_x, max_x, num=n_cells_x+1)
    y_steps = np.linspace(min_y, max_y, num=n_cells_y+1)

    goal_x_idx = np.max(np.where(x_steps <= x)[0])
    goal_y_idx = np.max(np.where(y_steps <= y)[0])
    return (goal_x_idx * n_cells_y) + goal_y_idx

timesteps = len(tj)
goals = np.zeros(timesteps)
for t in reversed(range(timesteps)):
    if (t + 1) % window == 0 or (t + 1) == timesteps:
        goals[t] = get_goal(tj[t], min_x, max_x, min_y, max_y, n_cells_x, n_cells_y)
        print('this:', t)
    else:
        goals[t] = goals[t + 1]
        

#%% check SBM datasets format
        
class DotDict(dict):
    def __init__(self, **kwds):
        self.update(kwds)
        self.__dict__ = self
        
import pickle

data_dir = 'data/SBMs/'
DATASET_NAME = 'SBM_CLUSTER'

with open(data_dir + DATASET_NAME + '.pkl', "rb") as f:
    unpickler = pickle.Unpickler(f)
    data = unpickler.load()
    # f = pickle.load(f)
    train = data[0]
    # test = data[1]
    # val = data[2]
#%%
from torch.utils.data import DataLoader

from data.SBMs import SBMsDataset, SBMsDatasetDGL
from data.TSP import TSPDataset
import dgl

class DotDict(dict):
    def __init__(self, **kwds):
        self.update(kwds)
        self.__dict__ = self
        
DATASET_NAME = 'SBM_CLUSTER'
dataset = SBMsDatasetDGL(DATASET_NAME)

dataset = SBMsDataset(DATASET_NAME) 
# dataset = TSPDataset(DATASET_NAME)
trainset, valset, testset = dataset.train, dataset.val, dataset.test

        
train_loader = DataLoader(trainset, batch_size=1, shuffle=False, collate_fn=dataset.collate)
val_loader = DataLoader(valset, batch_size=1, shuffle=False, collate_fn=dataset.collate)
test_loader = DataLoader(testset, batch_size=1, shuffle=False, collate_fn=dataset.collate)
    

for iter, (batch_graphs, batch_labels) in enumerate(train_loader):
    batch_x = batch_graphs.ndata['feat'].to("cuda")  # num x feat
    batch_e = batch_graphs.edata['feat'].to("cuda")
    batch_labels = batch_labels.to("cuda")
    break

batch_x_arry = batch_x.cpu().numpy()
batch_e_arry = batch_e.cpu().numpy()
batch_labels_arry = batch_labels.cpu().numpy()


#%% check tsp dataset
from scipy.spatial.distance import pdist, squareform
import numpy as np
import dgl
import torch

filename = './data/TSP/tsp50-500_test.txt'
file_data = open(filename, "r").readlines()[:100]
num_neighbors = 25

graph_lists = []
edge_labels = []
for graph_idx, line in enumerate(file_data):
    line = line.split(" ")  # Split into list
    num_nodes = int(line.index('output')//2)
    
    # Convert node coordinates to required format
    nodes_coord = []
    for idx in range(0, 2 * num_nodes, 2):
        nodes_coord.append([float(line[idx]), float(line[idx + 1])])

    # Compute distance matrix
    W_val = squareform(pdist(nodes_coord, metric='euclidean'))
    
    # Determine k-nearest neighbors for each node
    knns = np.argpartition(W_val, kth=num_neighbors, axis=-1)[:, num_neighbors::-1]

    # Convert tour nodes to required format
    # Don't add final connection for tour/cycle
    tour_nodes = [int(node) - 1 for node in line[line.index('output') + 1:-1]][:-1]

    # Compute an edge adjacency matrix representation of tour
    edges_target = np.zeros((num_nodes, num_nodes))
    for idx in range(len(tour_nodes) - 1):
        i = tour_nodes[idx]
        j = tour_nodes[idx + 1]
        edges_target[i][j] = 1
        edges_target[j][i] = 1
    # Add final connection of tour in edge target
    edges_target[j][tour_nodes[0]] = 1
    edges_target[tour_nodes[0]][j] = 1
    
    # Construct the DGL graph
    g = dgl.DGLGraph()
    g.add_nodes(num_nodes)
    g.ndata['feat'] = torch.Tensor(nodes_coord)
    
    edge_feats = []  # edge features i.e. euclidean distances between nodes
    edge_labels = []  # edges_targets as a list
    # Important!: order of edge_labels must be the same as the order of edges in DGLGraph g
    # We ensure this by adding them together
    for idx in range(num_nodes):
        for n_idx in knns[idx]:
            if n_idx != idx:  # No self-connection
                g.add_edge(idx, n_idx)
                edge_feats.append(W_val[idx][n_idx])
                edge_labels.append(int(edges_target[idx][n_idx]))
    # dgl.transform.remove_self_loop(g)
    
    # Sanity check
    assert len(edge_feats) == g.number_of_edges() == len(edge_labels)
    
    # Add edge features
    g.edata['feat'] = torch.Tensor(edge_feats).unsqueeze(-1)
    
    # # Uncomment to add dummy edge features instead (for Residual Gated ConvNet)
    # edge_feat_dim = g.ndata['feat'].shape[1] # dim same as node feature dim
    # g.edata['feat'] = torch.ones(g.number_of_edges(), edge_feat_dim)
    
    graph_lists.append(g)
    edge_labels.append(edge_labels)
    
    break

#%%
import os
from openpyxl import load_workbook, Workbook

run_dir = './out/GatedGCN/eth_trial_13'
run_dir_list = os.listdir(run_dir)

wb_dest = load_workbook('./out/GatedGCN/log_results.xlsx')
ws_dest = wb_dest.active

for run in run_dir_list:
    try:
        wb_source = load_workbook(os.path.join(run_dir, run, 'log_results.xlsx'))
        ws_source = wb_source.worksheets[0]  # select first worksheet
        row_data = [v for row in ws_source.iter_rows(min_row=2, max_col=20, max_row=2, values_only=True) for v in row]
        ws_dest.append(row_data)
        # break
    except FileNotFoundError:
        continue
wb_dest.save('./out/GatedGCN/log_results.xlsx')

#%%
#load arg.pkl
import pickle
out_dir = './out/SC_GCN/trial_17/'
run = 'run-5_hdim64_elayrs1_dlayrs2_zdim128_batch64_lr0.0001_kld0.1_edg0.0_mse1.0_cri0.0_gwt0.0_cwt0.0_center_scale_gcn'
run_dir = out_dir + run + '/'
with open(run_dir + 'args.pkl', 'rb') as f: 
    args = pickle.load(f)
    
#%%
#find min max in dataset
import os
import glob
import numpy as np
all_datasets = 'eth, hotel, univ, zara1, zara2'.split(', ')
data_stats = {}
all_data = []
for data_set in all_datasets:
    data_stats[data_set] = {}
    
    seq_data = []
    seq_data_rel = []
    for phase in ['test']:
        data_dir = './datasets/' + data_set + '/' + phase + '/'
        files_paths = glob.glob(data_dir + '*')
        
        for fpath in files_paths:
            if not '.txt' in fpath:
                continue
            data = np.loadtxt(fpath)
            print(data_set, len(np.unique(data[:, 0])))
            
            seq_data.append(data)    
            
    seq_data = np.concatenate(seq_data, axis=0)
    
    data_stats[data_set]['x_min'] = np.round(np.min(seq_data[:, 2]), decimals=3)
    data_stats[data_set]['x_max'] = np.round(np.max(seq_data[:, 2]), decimals=3)
    data_stats[data_set]['y_min'] = np.round(np.min(seq_data[:, 3]), decimals=3)
    data_stats[data_set]['y_max'] = np.round(np.max(seq_data[:, 3]), decimals=3)
    
    all_data.append(seq_data)
    
all_data = np.concatenate(all_data, axis=0)

data_stats['x_mean'] = np.round(np.mean(all_data[:, 2]), decimals=3)
data_stats['y_mean'] = np.round(np.mean(all_data[:, 3]), decimals=3)
data_stats['x_sigma'] = np.round(np.std(all_data[:, 2]), decimals=3)
data_stats['y_sigma'] = np.round(np.std(all_data[:, 3]), decimals=3)

        

#%%
# find number of ground truth collisions
import torch
from torch.utils.data import DataLoader
from utils import TrajectoryDataset
from config import parse_argument
from misc import setup_gpu
from train import organize_tensor, collect_collision_rewards

args = parse_argument()
device = setup_gpu(args.gpu_id, memory=args.gpu_memory)

gt_collision_count = {}
gt_collision_reward = {}
for args.dataset in 'eth, hotel, univ, zara1, zara2'.split(', '):
    print('Preparing datasets...', args.dataset)
    data_dir = './datasets/' + args.dataset + '/'

    phase = 'train'
    shuffle = False
    dataset = TrajectoryDataset(data_dir, obs_len=8, pred_len=12, phase=phase, preprocess=False)
    dataset._standardize_inputs(args.center, args.scale, args.data_format)
    dataloader = DataLoader(dataset, batch_size=args.batch_size, shuffle=shuffle, collate_fn=dataset.collate)
    gt_collisions = []
    gt_rewards = []
    for iter, (obsv_graph, target_graph) in enumerate(dataloader):
        # print("iter", iter)
        ped_id = obsv_graph.ndata['pid'].to(device) 
        
        target_p = target_graph.ndata['pos'].to(device) 
        target_f = target_graph.ndata['frames'].to(device)
        
        target_pos = organize_tensor(target_p, args.data_format, reshape=False, time_major=True)
        
        gt_collision_rewards, num_gt_collisions = collect_collision_rewards(target_pos, ped_id, target_f, args.collision_thresh)
        
        gt_collisions.append(num_gt_collisions)
        gt_rewards.append(torch.mean(gt_collision_rewards))
    
    gt_collision_count[args.dataset] = torch.stack(gt_collisions, dim=0).sum().cpu().item()
    gt_collision_reward[args.dataset] = torch.stack(gt_rewards, dim=0).mean().cpu().item()
    
#%% compare between linear and conv layer
def count_parameters(model):
    """Count the number of parameters in a model."""
    return sum([p.numel() for p in model.parameters()])

tensor = torch.randn(128, 2, 8, 1)
print(tensor.shape)

flattened_tensor = tensor.view(128, -1)
print(flattened_tensor.shape)

conv = torch.nn.Conv2d(2, 64, 1)
print(count_parameters(conv))
# 288

linear = torch.nn.Linear(16, 64)
print(count_parameters(linear))
# 288

# use same initialization
# linear.weight = torch.nn.Parameter(conv.weight.squeeze(2))
# linear.bias = torch.nn.Parameter(conv.bias)

out_linear = linear(flattened_tensor)
print(out_linear.shape)
# tensor(0.0067, grad_fn=<MeanBackward0>)

out_conv = conv(tensor)
print(out_conv.shape)
# tensor(0.0067, grad_fn=<MeanBackward0>)

# %timeit linear(flattened_tensor)
# 151 µs ± 297 ns per loop

# %timeit conv(tensor)
# 1.43 ms ± 6.33 µs per loop

#%% visualize edge weigths (eta_ij) of GatedGCN
res = model(obsv_graph, obsv_v, obsv_e, device=device)
g = res[0]
e_ij = g.edata['Ce'] + g.srcdata['Dh'] + g.dstdata['Eh']
sigma_ij = torch.sigmoid(e_ij)
Bh_j = g.srcdata['Bh']
eta_ij = torch.sum( sigma_ij * Bh_j, dim=1 ) / ( torch.sum( sigma_ij, dim=1 ) + 1e-6 )
print(eta_ij)

#%% collision detection
import math
import os
import numpy as np
from utils import read_file
total_collisions = []
total_peds_per_frame = []
avg_percent_collision_per_frame = []
percent_avg_collision_per_frame = []

collision_thresh = 0.2
for data_set in ['eth', 'hotel', 'univ', 'zara1', 'zara2']:
    data_dir = './datasets/' + data_set + '/' + 'test/'
    all_files = os.listdir(data_dir)
    all_files = [os.path.join(data_dir, _path) for _path in all_files]
    # print(all_files)
    
    for path in all_files:
        # print(path)
        seq_name = os.path.basename(path)[:-4]
        data = read_file(path, delim='\t')
        
        frames = np.unique(data[:, 0]).tolist()
        peds_per_frame = []
        collision_peds_per_frame = []
        for frame in frames:
            # frame_data.append(data[frame == data[:, 0], :])
            curr_frame_data = data[frame == data[:, 0], :]
            
            curr_num_peds = len(curr_frame_data)
            curr_collisions=0
            if curr_num_peds >1:
                for i in range(curr_num_peds ):
                    for j in range(i+1, curr_num_peds):
                        points1 = curr_frame_data[i, 2:4]
                        points2 = curr_frame_data[j, 2:4]
                        l2_dist = np.linalg.norm(points1 - points2)
                        if l2_dist<=collision_thresh:
                            curr_collisions+=1
                            
            collision_peds_per_frame.append(curr_collisions)
            peds_per_frame.append(curr_num_peds)
        
        # curr_avg_percent_collision_per_frame = np.mean([c/p for c, p in zip(collision_peds_per_frame, peds_per_frame)])*100
        curr_percent_of_avg_collision_per_frame = sum(collision_peds_per_frame)/sum(peds_per_frame)*100
        print('*'*10, seq_name, '*'*10)
        # print('Total collision peds per frame:', sum(collision_peds_per_frame))
        # print('Total peds per frame:', sum(peds_per_frame))
        # print('Average Percentage of collision peds per frame:', curr_avg_percent_collision_per_frame)
        print('Percentage of Avergae collision peds per frame', curr_percent_of_avg_collision_per_frame)
        
        # total_peds_per_frame.append(sum(peds_per_frame))
        # total_collisions.append(sum(collision_peds_per_frame))
        # avg_percent_collision_per_frame.append(curr_avg_percent_collision_per_frame)
        percent_avg_collision_per_frame.append(curr_percent_of_avg_collision_per_frame)
        
# print('Average peds per frame:', np.mean(total_peds_per_frame))
# print('Average collisions per frame', np.mean(total_collisions))
# print('Average percentage of collisions per frame:', np.mean(avg_percent_collision_per_frame))
print('Percentage of averaged collisions per frame:', np.mean(percent_avg_collision_per_frame))
my_collision = [0.03, 0.0, 0.005, 0.33, 0.0, 0.041]
sophie_collision_peds = [1.757, 1.936, 0.621, 1.027, 1.464]

#%%

import math
import os
import numpy as np
from utils import read_file

avg_collision_per_frame = []
collision_thresh = 0.2
for data_set in ['eth', 'hotel', 'univ', 'zara1', 'zara2']:
    data_dir = './datasets/' + data_set + '/' + 'test/'
    all_files = os.listdir(data_dir)
    all_files = [os.path.join(data_dir, _path) for _path in all_files]
    # print(all_files)
    
    for path in all_files:
        # print(path)
        seq_name = os.path.basename(path)[:-4]
        data = read_file(path, delim='\t')
        
        frames = np.unique(data[:, 0]).tolist()
        peds_per_frame = []
        collision_peds_per_frame = []
        for frame in frames:
            # frame_data.append(data[frame == data[:, 0], :])
            curr_frame_data = data[frame == data[:, 0], :]
            
            curr_num_peds = len(curr_frame_data)
            curr_collisions=0
            if curr_num_peds >1:
                for i in range(curr_num_peds ):
                    for j in range(i+1, curr_num_peds):
                        points1 = curr_frame_data[i, 2:4]
                        points2 = curr_frame_data[j, 2:4]
                        l2_dist = np.linalg.norm(points1 - points2)
                        if l2_dist<=collision_thresh:
                            curr_collisions+=1
                            
            collision_peds_per_frame.append(curr_collisions)
            peds_per_frame.append(curr_num_peds)
        
        avg_percent_collision_per_frame = sum(collision_peds_per_frame)/sum(peds_per_frame)*100
        print('*'*10, seq_name, '*'*10)

        print('Percentage of Avergae collision peds per frame', avg_percent_collision_per_frame)

        avg_collision_per_frame.append(avg_percent_collision_per_frame)

print('Percentage of averaged collisions per frame:', np.mean(avg_collision_per_frame))
#my_collision = [0.03, 0.0, 0.005, 0.33, 0.0, 0.041]
#sophie_collision_peds = [1.757, 1.936, 0.621, 1.027, 1.464]